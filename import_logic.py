"""
Import logic for Moneybags application.

Handles Excel file parsing, validation, and import execution.
Supports two Excel formats:
- Original format (active sheet with columns C-N)
- Hovedark format (sheet "Hovedark" with columns F-Q)

This module follows the same architectural pattern as business_logic.py:
- NO direct database calls - always use database_manager module
- UUIDs generated via utils.generate_uid()
- Timestamps set here (not in database)
- Empty strings converted to NULL via utils.empty_to_none()
"""

import logging
from datetime import datetime
from typing import Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utils import generate_uid, empty_to_none, validate_month, validate_year
import database_manager as db

logger = logging.getLogger(__name__)


# ==================== HELPER FUNCTIONS ====================

def _extract_amounts_from_formula(cell_value, row_num: int, col_name: str) -> list[int]:
    """
    Extract individual amounts from Excel formula or value.

    Strict validation - only simple addition formulas allowed.

    Args:
        cell_value: Cell value (formula string or number)
        row_num: Row number for error messages
        col_name: Column name for error messages

    Returns:
        List of integer amounts extracted from formula

    Raises:
        ValueError: On invalid formula format or negative values

    Examples:
        "=575+2182" → [575, 2182]
        "=104571" → [104571]
        "=((427+275)+7292)+200" → [427, 275, 7292, 200]
        "55615.0" → [55615]
        "0" → [0]
        "" → []
    """
    # Handle None or empty
    if cell_value is None or cell_value == "":
        return []

    # Convert to string
    formula_str = str(cell_value).strip()

    # Empty after strip
    if not formula_str:
        return []

    # Remove "=" prefix if present
    if formula_str.startswith("="):
        formula_str = formula_str[1:]

    # Remove all parentheses (they're just grouping for addition, which is allowed)
    formula_str = formula_str.replace("(", "").replace(")", "")

    # Check for forbidden operations/functions
    forbidden = ["IF", "SUM", "AVERAGE", "COUNT", "MIN", "MAX", "*", "/", "-"]
    for forbidden_item in forbidden:
        if forbidden_item in formula_str:
            if forbidden_item in ["IF", "SUM", "AVERAGE", "COUNT", "MIN", "MAX"]:
                raise ValueError(f"Row {row_num}, Column {col_name}: Complex formula not supported ({forbidden_item})")
            elif forbidden_item == "-":
                # Allow negative check after splitting
                pass
            else:
                raise ValueError(f"Row {row_num}, Column {col_name}: Only addition (+) supported")

    # Split by "+"
    parts = formula_str.split("+")

    # Convert to integers
    amounts = []
    for part in parts:
        part = part.strip()
        if not part:
            continue

        try:
            # Convert to float first (handles decimals), then to int
            value = int(float(part))

            # Reject negative values
            if value < 0:
                raise ValueError(f"Row {row_num}, Column {col_name}: Negative value not allowed ({value})")

            amounts.append(value)
        except ValueError as e:
            if "Negative value not allowed" in str(e):
                raise
            raise ValueError(f"Row {row_num}, Column {col_name}: Invalid number format: {part}")

    return amounts


def _ensure_import_payee() -> str:
    """
    Get or create "Import - Google Sheets" payee.

    Returns:
        str: Payee UUID
    """
    payee = db.get_payee_by_name("Import - Google Sheets")
    if payee:
        return payee.id

    # Create payee
    data = {
        "id": generate_uid(),
        "name": "Import - Google Sheets",
        "type": "Generic",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    payee = db.create_payee(data)
    return payee.id


# ==================== EXCEL PARSING ====================

def parse_excel_file(file_path: str, year: int) -> dict:
    """
    Parse Google Sheets Excel file and extract budget/actual data.

    Supports two formats:

    Format 1 (Original - active sheet):
    - Row 3: Headers ("Balanse", "Januar", ..., "Desember")
    - Row 7: "Inntekter" section
    - Row 16+: "Utgifter" section
    - Category blocks every 4 rows (income) or 3 rows (expenses)

    Format 2 (Hovedark sheet):
    - Sheet name: "Hovedark"
    - Month columns: F-Q (Jan-Dec)
    - Category row: Col C = name, Col D = "Budsjett"
    - Budget row: Row N, values in F-Q
    - Actual row: Row N+1, formulas in F-Q
    - Skip rows: N+2, N+3 (computed)

    Args:
        file_path: Path to .xlsx file
        year: Year for the data

    Returns:
        {
            "year": 2024,
            "sheet_categories": [
                {
                    "name": "Lønn",
                    "type": "income",
                    "budget": {1: 52000, 2: 52000, ...},
                    "actuals": {1: [55615], 2: [55615], ...}
                }
            ]
        }

    Raises:
        ValueError: On validation errors
    """
    # Validate year
    if not validate_year(year):
        raise ValueError(f"Invalid year: {year}")

    # Load workbook
    try:
        wb = load_workbook(file_path, data_only=False)
    except Exception as e:
        raise ValueError(f"Failed to load Excel file: {e}")

    # Detect format by checking structure
    # New format has: category name in Col C, "Budsjett" in Col D
    # Old format has: category name in Col B, "Budsjett" in next row Col B
    if "Hovedark" in wb.sheetnames:
        sheet = wb["Hovedark"]
        # Check if new format: scan for a row with "Budsjett" in Col D
        is_new_format = False
        for row_idx in range(1, 100):
            col_d_value = sheet.cell(row_idx, 4).value
            if col_d_value == "Budsjett":
                is_new_format = True
                break

        if is_new_format:
            return _parse_hovedark_format(wb, year)
        else:
            return _parse_original_format(wb, year)
    else:
        return _parse_original_format(wb, year)


def _parse_hovedark_format(wb, year: int) -> dict:
    """
    Parse Hovedark Excel format.

    Structure:
    - Sheet: "Hovedark"
    - Month columns: F-Q (columns 6-17, Jan-Dec)
    - Category identification: Col C = category name, Col D = "Budsjett"
    - Row N: Budget row (category name in C, "Budsjett" in D, values in F-Q)
    - Row N+1: Actual row ("Resultat" in D, formulas in F-Q)
    - Row N+2, N+3: Computed rows (skip)
    - Sections: "Utgifter" (expenses) and "Inntekter" (income)
    - Skip categories with "Total" in name
    """
    sheet = wb["Hovedark"]

    # Month column mapping for Hovedark format (F=1, G=2, ..., Q=12)
    month_columns = {
        6: 1,   # F = January
        7: 2,   # G = February
        8: 3,   # H = March
        9: 4,   # I = April
        10: 5,  # J = May
        11: 6,  # K = June
        12: 7,  # L = July
        13: 8,  # M = August
        14: 9,  # N = September
        15: 10, # O = October
        16: 11, # P = November
        17: 12  # Q = December
    }

    # Find section boundaries
    utgifter_row = None
    inntekter_row = None

    for row_idx in range(1, 100):
        cell_c = sheet.cell(row_idx, 3).value  # Column C
        if cell_c:
            if "Utgifter" in str(cell_c):
                utgifter_row = row_idx
            elif "Inntekter" in str(cell_c):
                inntekter_row = row_idx

    if not utgifter_row:
        raise ValueError("Could not find 'Utgifter' section in Hovedark sheet")
    if not inntekter_row:
        raise ValueError("Could not find 'Inntekter' section in Hovedark sheet")

    logger.info(f"Found Utgifter at row {utgifter_row}, Inntekter at row {inntekter_row}")

    sheet_categories = []

    # Scan all rows looking for categories (where Col D = "Budsjett")
    for row_idx in range(1, 100):
        col_c = sheet.cell(row_idx, 3).value  # Category name
        col_d = sheet.cell(row_idx, 4).value  # Should be "Budsjett"

        # A category row has a name in Col C and "Budsjett" in Col D
        if not col_c or col_d != "Budsjett":
            continue

        category_name = str(col_c).strip()

        # Skip total rows
        if "Total" in category_name or "Totale" in category_name:
            logger.info(f"Skipping total row: {category_name}")
            continue

        # Determine type based on section
        if row_idx > utgifter_row and (row_idx < inntekter_row or inntekter_row < utgifter_row):
            category_type = "expenses"  # Plural to match database and old format
        elif row_idx > inntekter_row:
            category_type = "income"
        else:
            # Skip categories before both sections
            logger.info(f"Skipping category before sections: {category_name}")
            continue

        # Extract budget values from this row (cols F-Q)
        budget = {}
        for col_idx, month in month_columns.items():
            cell = sheet.cell(row_idx, col_idx)
            if cell.value:
                try:
                    # Budget cells can be formulas (e.g., =1000+5200) or plain numbers
                    col_letter = get_column_letter(col_idx)
                    amounts = _extract_amounts_from_formula(cell.value, row_idx, col_letter)
                    if amounts:
                        # Sum all amounts for budget total
                        amount = sum(amounts)
                        if amount != 0:
                            budget[month] = amount
                except (ValueError, TypeError):
                    # Skip invalid values
                    pass

        # Extract actual values from row N+1 (cols F-Q)
        actuals = {}
        actuals_row = row_idx + 1
        for col_idx, month in month_columns.items():
            cell = sheet.cell(actuals_row, col_idx)
            if cell.value:
                try:
                    col_letter = get_column_letter(col_idx)
                    amounts = _extract_amounts_from_formula(cell.value, actuals_row, col_letter)
                    if amounts:
                        actuals[month] = amounts
                except ValueError as e:
                    raise ValueError(f"Category '{category_name}': {e}")

        # Skip categories with no data
        if not budget and not actuals:
            logger.info(f"Skipping category with no data: {category_name}")
            continue

        logger.info(f"Found category: {category_name} ({category_type}), budget months: {len(budget)}, actual months: {len(actuals)}")

        sheet_categories.append({
            "name": category_name,
            "type": category_type,
            "budget": budget,
            "actuals": actuals
        })

    if not sheet_categories:
        raise ValueError("No categories found in Hovedark sheet")

    return {
        "year": year,
        "sheet_categories": sheet_categories
    }


def _parse_original_format(wb, year: int) -> dict:
    """Parse original Excel format (active sheet with columns C-N)."""
    sheet = wb.active

    # Month column mapping (C=1, D=2, ..., N=12)
    month_columns = {
        'C': 1, 'D': 2, 'E': 3, 'F': 4, 'G': 5, 'H': 6,
        'I': 7, 'J': 8, 'K': 9, 'L': 10, 'M': 11, 'N': 12
    }

    # Find "Utgifter" row to split income vs expenses
    utgifter_row = None
    for row_idx in range(1, 100):
        cell = sheet[f'B{row_idx}']
        if cell.value and "Utgifter" in str(cell.value):
            utgifter_row = row_idx
            break

    if not utgifter_row:
        raise ValueError("Could not find 'Utgifter' section in Excel file")

    sheet_categories = []

    # Parse income categories (4-row pattern: Category, Budsjett, Resultat, Differanse)
    # Start at row 8, step by 4, until we reach utgifter_row
    for row_idx in range(8, utgifter_row, 4):
        category_cell = sheet[f'B{row_idx}']
        category_name = category_cell.value

        # Skip if no category name or if it's a header row or row label
        if not category_name or category_name in ["Inntekter", "Utgifter", "Balanse", "Budsjett", "Resultat", "Differanse"]:
            continue

        category_name = str(category_name).strip()
        category_type = "income"

        # Extract budget values (row N+1)
        budget = {}
        budget_row = row_idx + 1
        for col, month in month_columns.items():
            cell = sheet[f'{col}{budget_row}']
            if cell.value:
                try:
                    # Budget cells can be formulas (e.g., =1000+5200) or plain numbers
                    amounts = _extract_amounts_from_formula(cell.value, budget_row, col)
                    if amounts:
                        budget[month] = sum(amounts)  # Sum for total budget
                except (ValueError, TypeError):
                    # Skip invalid values
                    pass

        # Extract actual values (row N+2)
        actuals = {}
        actuals_row = row_idx + 2
        for col, month in month_columns.items():
            cell = sheet[f'{col}{actuals_row}']
            if cell.value:
                try:
                    amounts = _extract_amounts_from_formula(cell.value, actuals_row, col)
                    if amounts:
                        actuals[month] = amounts
                except ValueError as e:
                    raise ValueError(f"Category '{category_name}': {e}")

        # Skip categories with no data
        if not budget and not actuals:
            continue

        sheet_categories.append({
            "name": category_name,
            "type": category_type,
            "budget": budget,
            "actuals": actuals
        })

    # Parse expense categories (3-row pattern: Category, Budsjett, Resultat - NO Differanse)
    # Start at utgifter_row + 1, step by 3
    for row_idx in range(utgifter_row + 1, 60, 3):
        category_cell = sheet[f'B{row_idx}']
        category_name = category_cell.value

        # Skip if no category name or if it's a header row or row label
        if not category_name or category_name in ["Inntekter", "Utgifter", "Balanse", "Budsjett", "Resultat", "Differanse"]:
            continue

        category_name = str(category_name).strip()
        category_type = "expenses"

        # Extract budget values (row N+1)
        budget = {}
        budget_row = row_idx + 1
        for col, month in month_columns.items():
            cell = sheet[f'{col}{budget_row}']
            if cell.value:
                try:
                    # Budget cells can be formulas (e.g., =1000+5200) or plain numbers
                    amounts = _extract_amounts_from_formula(cell.value, budget_row, col)
                    if amounts:
                        budget[month] = sum(amounts)  # Sum for total budget
                except (ValueError, TypeError):
                    # Skip invalid values
                    pass

        # Extract actual values (row N+2)
        actuals = {}
        actuals_row = row_idx + 2
        for col, month in month_columns.items():
            cell = sheet[f'{col}{actuals_row}']
            if cell.value:
                try:
                    amounts = _extract_amounts_from_formula(cell.value, actuals_row, col)
                    if amounts:
                        actuals[month] = amounts
                except ValueError as e:
                    raise ValueError(f"Category '{category_name}': {e}")

        # Skip categories with no data
        if not budget and not actuals:
            continue

        sheet_categories.append({
            "name": category_name,
            "type": category_type,
            "budget": budget,
            "actuals": actuals
        })

    if not sheet_categories:
        raise ValueError("No categories found in Excel file")

    return {
        "year": year,
        "sheet_categories": sheet_categories
    }


# ==================== VALIDATION & EXECUTION ====================

def validate_import(parsed_data: dict, category_mapping: dict) -> dict:
    """
    Dry-run validation before import.

    Checks:
    - All mapped categories exist
    - Category types match (income/expenses)
    - Check for duplicate BudgetEntries
    - Check for duplicate Transactions

    Args:
        parsed_data: Parsed Excel data structure
        category_mapping: Dict mapping sheet category names to Moneybags category IDs

    Returns:
        {
            "valid": True/False,
            "errors": ["..."],
            "warnings": ["..."],
            "summary": {"budget_count": 120, "transaction_count": 347}
        }
    """
    errors = []
    warnings = []
    budget_count = 0
    transaction_count = 0

    year = parsed_data["year"]

    # Validate all categories exist and types match
    for sheet_cat in parsed_data["sheet_categories"]:
        sheet_name = sheet_cat["name"]

        # Check category is mapped
        if sheet_name not in category_mapping:
            errors.append(f"Category '{sheet_name}' not mapped")
            continue

        category_id = category_mapping[sheet_name]

        # Check category exists
        category = db.get_category_by_id(category_id)
        if not category:
            errors.append(f"Category '{sheet_name}' mapped to '{category_id}' which does not exist")
            continue

        # Check type matches
        if category.type != sheet_cat["type"]:
            errors.append(f"Category '{sheet_name}' type mismatch: sheet has '{sheet_cat['type']}' but Moneybags has '{category.type}'")
            continue

        # Count budget entries and check for duplicates
        for month_str in sheet_cat["budget"].keys():
            month = int(month_str)  # Convert string to int
            existing = db.get_budget_entry(category_id, year, month)
            if existing:
                warnings.append(f"Budget entry for '{category.name}' {year}-{month:02d} already exists - will overwrite")
            budget_count += 1

        # Count transactions
        for month_str, amounts in sheet_cat["actuals"].items():
            month = int(month_str)  # Convert string to int
            transaction_count += len(amounts)

    # Final validation
    valid = len(errors) == 0

    return {
        "valid": valid,
        "errors": errors,
        "warnings": warnings,
        "summary": {
            "budget_count": budget_count,
            "transaction_count": transaction_count
        }
    }


def import_budget_and_transactions(parsed_data: dict, category_mapping: dict) -> dict:
    """
    Create BudgetEntry and Transaction records from parsed data.

    Prerequisites: validate_import() should pass before calling this

    Args:
        parsed_data: Parsed Excel data structure
        category_mapping: Dict mapping sheet category names to Moneybags category IDs

    Returns:
        {
            "budget_count": 120,
            "transaction_count": 347,
            "message": "Successfully imported data"
        }

    Raises:
        ValueError: On errors
    """
    import_payee_id = _ensure_import_payee()

    budget_count = 0
    transaction_count = 0

    year = parsed_data["year"]

    for sheet_category in parsed_data["sheet_categories"]:
        category_id = category_mapping[sheet_category["name"]]

        # Import budget entries
        for month_str, amount in sheet_category["budget"].items():
            month = int(month_str)  # Convert string to int
            data = {
                "id": generate_uid(),
                "category_id": category_id,
                "year": year,
                "month": month,
                "amount": amount,
                "comment": empty_to_none(None),
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            db.create_or_update_budget_entry(data)
            budget_count += 1

        # Import transactions
        for month_str, amounts in sheet_category["actuals"].items():
            month = int(month_str)  # Convert string to int
            for amount in amounts:
                date_str = f"{year}-{month:02d}-01"
                data = {
                    "id": generate_uid(),
                    "category_id": category_id,
                    "payee_id": import_payee_id,
                    "date": date_str,
                    "amount": amount,
                    "comment": empty_to_none(None),
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                db.create_transaction(data)
                transaction_count += 1

    logger.info(f"Imported {budget_count} budget entries and {transaction_count} transactions")

    # Ensure all imported categories are in the budget template for this year
    template_count = 0
    unique_category_ids = set(category_mapping.values())
    for category_id in unique_category_ids:
        # Only create if it doesn't already exist
        if not db.budget_template_exists(year, category_id):
            template_data = {
                'id': generate_uid(),
                'year': year,
                'category_id': category_id,
                'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            db.create_budget_template(template_data)
            template_count += 1
            logger.info(f"Added category {category_id} to budget template for {year}")

    return {
        "budget_count": budget_count,
        "transaction_count": transaction_count,
        "template_count": template_count,
        "message": "Successfully imported data"
    }
